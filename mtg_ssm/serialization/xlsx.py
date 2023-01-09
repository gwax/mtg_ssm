"""XLSX serializer."""

import collections
import datetime as dt
import string
from pathlib import Path
from typing import Any, ClassVar, Dict, Iterable, List, Optional, Sequence, Set, Tuple

import openpyxl
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from mtg_ssm.containers import counts
from mtg_ssm.containers.collection import MagicCollection
from mtg_ssm.containers.indexes import Oracle, ScryfallDataIndex
from mtg_ssm.mtg import util
from mtg_ssm.scryfall.models import ScryCard, ScrySet
from mtg_ssm.serialization import interface

ALL_SETS_SHEET_HEADER: Sequence[str] = [
    "code",
    "name",
    "release",
    "block",
    "type",
    "cards",
    "unique",
    "playsets",
    "count",
    "value",
    "nonbulk",
]

ALL_SETS_SHEET_TOTALS: Sequence[Optional[str]] = ["Total", None, None, None, None] + [
    f"=SUM({c}3:{c}65535)" for c in "FGHIJK"
]


def _card_set_sort_key(cset: ScrySet) -> Tuple[dt.date, str, int, str]:
    released_at = cset.released_at or dt.date.min
    return (
        released_at,
        cset.parent_set_code or cset.code,
        0 if cset.parent_set_code is None else 1,
        cset.code,
    )


def create_all_sets(sheet: Worksheet, index: ScryfallDataIndex) -> None:
    """Create all sets sheet from card_db."""
    sheet.title = "All Sets"
    sheet.append(ALL_SETS_SHEET_HEADER)
    sheet.append(ALL_SETS_SHEET_TOTALS)
    for card_set in sorted(index.setcode_to_set.values(), key=_card_set_sort_key):
        setcode = card_set.code.upper()
        row = [
            setcode,
            card_set.name,
            card_set.released_at,
            card_set.block,
            card_set.set_type.value,
            len(index.setcode_to_cards[card_set.code]),
            f"=COUNTIF('{setcode}'!{_setsheet_col('have')}:{_setsheet_col('have')},\">0\")",
            f"=COUNTIF('{setcode}'!{_setsheet_col('have')}:{_setsheet_col('have')},\">=4\")",
            f"=SUM('{setcode}'!{_setsheet_col('have')}:{_setsheet_col('have')})",
            f"=SUM('{setcode}'!{_setsheet_col('value')}:{_setsheet_col('value')})",
            f"=SUMIFS('{setcode}'!{_setsheet_col('value')}:{_setsheet_col('value')},'{setcode}'!{_setsheet_col('price')}:{_setsheet_col('price')},\">=1\")",
        ]
        sheet.append(row)


def style_all_sets(sheet: Worksheet) -> None:
    """Apply styles to the all sets sheet."""
    sheet.freeze_panes = sheet["C3"]
    col_width_hidden_format = [
        ("A", 8, False, None),
        ("B", 30, False, None),
        ("C", 12, True, None),
        ("D", 22, True, None),
        ("E", 15, True, None),
        ("F", 6, False, None),
        ("G", 7, False, None),
        ("H", 8, False, None),
        ("I", 7, False, None),
        ("J", 10, False, FORMAT_CURRENCY_USD_SIMPLE),
        ("K", 10, False, FORMAT_CURRENCY_USD_SIMPLE),
    ]
    for col, width, hidden, number_format in col_width_hidden_format:
        cdim = sheet.column_dimensions[col]
        cdim.width = width
        cdim.hidden = hidden
        if number_format is not None:
            cdim.number_format = number_format


def create_haverefs(index: ScryfallDataIndex, cards: Sequence[ScryCard]) -> str:
    """Create a reference to the have cells for printings in a single set."""
    setcodes_and_rownums = sorted(
        (c.set, index.id_to_setindex[c.id] + ROW_OFFSET) for c in cards
    )
    haverefs = [
        f"'{setcode.upper()}'!A{rownum}" for setcode, rownum in setcodes_and_rownums
    ]
    return "+".join(haverefs)


def get_references(
    index: ScryfallDataIndex, card_name: str, exclude_sets: Optional[Set[str]] = None
) -> Optional[str]:
    """Get an equation for the references to a card."""
    if util.is_strict_basic(card_name):
        return None  # Basics are so prolific that they overwhelm Excel

    if exclude_sets is None:
        exclude_sets = set()

    set_to_cards: Dict[str, List[ScryCard]] = collections.defaultdict(list)
    for other_card in index.name_to_cards[card_name]:
        if other_card.set not in exclude_sets:
            set_to_cards[other_card.set].append(other_card)

    set_to_haveref = {k: create_haverefs(index, v) for k, v in set_to_cards.items()}
    if not set_to_haveref:
        return None

    references = []
    for card_set in sorted(
        set_to_haveref,
        key=lambda setcode: _card_set_sort_key(index.setcode_to_set[setcode]),
    ):
        reference = 'IF({count}>0,"{setcode}: "&{count}&", ","")'.format(
            setcode=card_set.upper(), count=set_to_haveref[card_set]
        )
        references.append(reference)
    return "=" + "&".join(references)


ALL_CARDS_SHEET_HEADER = ["name", "have"]  # TODO: add list of sets


def create_all_cards(sheet: Worksheet, index: ScryfallDataIndex) -> None:
    """Create all cards sheet from card_db."""
    sheet.title = "All Cards"
    sheet.append(ALL_CARDS_SHEET_HEADER)
    for name in sorted(index.name_to_cards):
        row = [name, get_references(index, name)]
        sheet.append(row)


def style_all_cards(sheet: Worksheet) -> None:
    """Apply styles to the all cards sheet."""
    sheet.freeze_panes = sheet["B2"]
    col_width_hidden = [("A", 24, False), ("B", 32, False)]
    for col, width, hidden in col_width_hidden:
        cdim = sheet.column_dimensions[col]
        cdim.width = width
        cdim.hidden = hidden


SET_SHEET_HEADER = (
    [
        "have",
        "value",
        "name",
        "number",
        "scryfall_id",
        "artist",
        "price",
        "foil_price",
    ]
    + [ct.value for ct in counts.CountType]
    + ["others"]
)


def _setsheet_col(column_header: str) -> str:
    return string.ascii_uppercase[SET_SHEET_HEADER.index(column_header)]


COUNT_COLS = [_setsheet_col(ct) for ct in counts.CountType]
HAVE_TMPL = "=" + "+".join(col + "{rownum}" for col in COUNT_COLS)
HAVE_TMPL = (
    "="
    + _setsheet_col(counts.CountType.NONFOIL)
    + "{rownum}"
    + "+"
    + _setsheet_col(counts.CountType.FOIL)
    + "{rownum}"
)
VALUE_TMPL = (
    "="
    + _setsheet_col(counts.CountType.NONFOIL)
    + "{rownum}"
    + "*"
    + _setsheet_col("price")
    + "{rownum}"
    + "+"
    + _setsheet_col(counts.CountType.FOIL)
    + "{rownum}"
    + "*"
    + _setsheet_col("foil_price")
    + "{rownum}"
)
ROW_OFFSET = 2


def create_set_sheet(
    sheet: Worksheet, collection: MagicCollection, setcode: str
) -> None:
    """Populate sheet with card information from a given set."""
    index = collection.oracle.index

    sheet.append(SET_SHEET_HEADER)
    sheet.title = setcode.upper()

    for card in index.setcode_to_cards[setcode]:
        rownum = ROW_OFFSET + index.id_to_setindex[card.id]
        row: List[Optional[Any]] = [
            HAVE_TMPL.format(rownum=rownum),
            VALUE_TMPL.format(rownum=rownum),
            card.name,
            card.collector_number,
            str(card.id),
            card.artist,
            (card.prices or {}).get("usd", None),
            (card.prices or {}).get("usd_foil", None),
        ]
        card_counts = collection.counts.get(card.id, {})
        for count_type in counts.CountType:
            row.append(card_counts.get(count_type))
        row.append(get_references(index, card.name, exclude_sets={setcode}))
        sheet.append(row)


def style_set_sheet(sheet: Worksheet) -> None:
    """Apply styles to a set sheet."""
    sheet.freeze_panes = sheet["E2"]
    col_width_hidden_format = [
        ("A", 5, False, None),
        ("B", 9, False, FORMAT_CURRENCY_USD_SIMPLE),
        ("C", 24, False, None),
        ("D", 7, False, None),
        ("E", 10, True, None),
        ("F", 20, True, None),
        ("G", 9, True, FORMAT_CURRENCY_USD_SIMPLE),
        ("H", 9, True, FORMAT_CURRENCY_USD_SIMPLE),
        ("I", 7, False, None),
        ("J", 6, False, None),
        ("K", 10, False, None),
    ]
    for col, width, hidden, number_format in col_width_hidden_format:
        cdim = sheet.column_dimensions[col]
        cdim.width = width
        cdim.hidden = hidden
        if number_format is not None:
            cdim.number_format = number_format


def rows_from_sheet(sheet: Worksheet) -> Iterable[Dict[str, str]]:
    """Read rows from an xlsx worksheet as dicts."""
    header_row, *rows = iter(sheet.rows)
    header = [cell.value for cell in header_row]
    for row in rows:
        values = [cell.value for cell in row]
        if any(v is not None for v in values):
            yield dict(zip(header, values), set=str(sheet.title))


def rows_for_workbook(
    book: Workbook, *, skip_sheets: Optional[Set[str]]
) -> Iterable[Dict[str, str]]:
    """Read rows from an xlsx workbook as dicts."""
    if skip_sheets is None:
        skip_sheets = set()
    for sheet in book.worksheets:
        if sheet.title in skip_sheets:
            continue
        yield from rows_from_sheet(sheet)


class XlsxDialect(interface.SerializationDialect):
    """excel xlsx collection"""

    extension: ClassVar[str] = "xlsx"
    dialect: ClassVar[str] = "xlsx"

    def write(self, path: Path, collection: MagicCollection) -> None:
        """Write collection to an xlsx file."""
        workbook = openpyxl.Workbook()

        all_sets_sheet = workbook.create_sheet()
        create_all_sets(all_sets_sheet, collection.oracle.index)
        style_all_sets(all_sets_sheet)

        all_cards_sheet = workbook.create_sheet()
        create_all_cards(all_cards_sheet, collection.oracle.index)
        style_all_cards(all_cards_sheet)

        setcodes = [
            s.code
            for s in sorted(
                collection.oracle.index.setcode_to_set.values(),
                key=_card_set_sort_key,
            )
        ]

        for setcode in setcodes:
            set_sheet = workbook.create_sheet()
            create_set_sheet(set_sheet, collection, setcode)
            style_set_sheet(set_sheet)
        del workbook["Sheet"]
        workbook.save(str(path))

    def read(self, path: Path, oracle: Oracle) -> MagicCollection:
        """Read collection from an xlsx file."""
        workbook = openpyxl.load_workbook(filename=str(path), read_only=True)
        reader = rows_for_workbook(workbook, skip_sheets={"All Sets", "All Cards"})
        card_counts = counts.aggregate_card_counts(reader, oracle)
        return MagicCollection(oracle=oracle, counts=card_counts)
