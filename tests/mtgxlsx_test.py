"""Tests for mtgcdb.mtgxlsx"""

import collections
import datetime
from unittest import mock

import openpyxl

from mtgcdb import models
from mtgcdb import mtgjson
from mtgcdb import mtgxlsx

from tests import mtgjson_testcase
from tests import sqlite_testcase


class MtgXlsxTest(
        sqlite_testcase.SqliteTestCase, mtgjson_testcase.MtgJsonTestCase):

    def setUp(self):
        super().setUp()
        connection = self.engine.connect()
        models.Base.metadata.create_all(connection)
        connection.close()

    def test_create_sets_sheet(self):
        # Setup
        mtgjson.update_models(self.session, self.mtg_data, True)
        self.session.commit()
        card_sets = self.session.query(models.CardSet) \
            .order_by(models.CardSet.release_date).all()
        book = openpyxl.Workbook()
        sheet = book.create_sheet()

        # Execute
        mtgxlsx.create_sets_sheet(sheet, card_sets)

        # Verify
        rows = [[cell.value for cell in row] for row in sheet.rows]
        # pylint: disable=line-too-long
        expected = [
            ['code', 'name', 'release', 'block', 'type', 'cards', 'unique', 'playsets', 'count'],
            ['LEA', 'Limited Edition Alpha', datetime.datetime(1993, 8, 5), None, 'core', 4, '=COUNTIF(\'LEA\'!A:A,">0")', '=COUNTIF(\'LEA\'!A:A,">=4")', "=SUM('LEA'!A:A)"],
            ['ICE', 'Ice Age', datetime.datetime(1995, 6, 1), 'Ice Age', 'expansion', 5, '=COUNTIF(\'ICE\'!A:A,">0")', '=COUNTIF(\'ICE\'!A:A,">=4")', "=SUM('ICE'!A:A)"],
            ['HML', 'Homelands', datetime.datetime(1995, 10, 1), None, 'expansion', 2, '=COUNTIF(\'HML\'!A:A,">0")', '=COUNTIF(\'HML\'!A:A,">=4")', "=SUM('HML'!A:A)"],
            ['S00', 'Starter 2000', datetime.datetime(2000, 4, 1), None, 'starter', 1, '=COUNTIF(\'S00\'!A:A,">0")', '=COUNTIF(\'S00\'!A:A,">=4")', "=SUM('S00'!A:A)"],
            ['pMGD', 'Magic Game Day', datetime.datetime(2007, 7, 14), None, 'promo', 1, '=COUNTIF(\'pMGD\'!A:A,">0")', '=COUNTIF(\'pMGD\'!A:A,">=4")', "=SUM('pMGD'!A:A)"],
            ['HOP', 'Planechase', datetime.datetime(2009, 9, 4), None, 'planechase', 3, '=COUNTIF(\'HOP\'!A:A,">0")', '=COUNTIF(\'HOP\'!A:A,">=4")', "=SUM('HOP'!A:A)"],
            ['ARC', 'Archenemy', datetime.datetime(2010, 6, 18), None, 'archenemy', 2, '=COUNTIF(\'ARC\'!A:A,">0")', '=COUNTIF(\'ARC\'!A:A,">=4")', "=SUM('ARC'!A:A)"],
            ['ISD', 'Innistrad', datetime.datetime(2011, 9, 30), 'Innistrad', 'expansion', 6, '=COUNTIF(\'ISD\'!A:A,">0")', '=COUNTIF(\'ISD\'!A:A,">=4")', "=SUM('ISD'!A:A)"],
            ['PC2', 'Planechase 2012 Edition', datetime.datetime(2012, 6, 1), None, 'planechase', 4, '=COUNTIF(\'PC2\'!A:A,">0")', '=COUNTIF(\'PC2\'!A:A,">=4")', "=SUM('PC2'!A:A)"],
            ['VMA', 'Vintage Masters', datetime.datetime(2014, 6, 16), None, 'masters', 1, '=COUNTIF(\'VMA\'!A:A,">0")', '=COUNTIF(\'VMA\'!A:A,">=4")', "=SUM('VMA'!A:A)"],
        ]
        # pylint: enable=line-too-long
        self.assertEqual(expected, rows)

    def test_get_other_print_references(self):
        # Setup
        mtgjson.update_models(self.session, self.mtg_data, True)
        self.session.commit()
        dark_rits = self.session.query(models.CardPrinting).filter(
            models.CardPrinting.card.has(name='Dark Ritual')).all()
        name_to_prints = {'Dark Ritual': dark_rits}
        lea_dark_rit = self.session.query(
            models.CardPrinting).filter_by(multiverseid=54).first()

        # Execute
        print_refs = mtgxlsx.get_other_print_references(
            lea_dark_rit, name_to_prints)

        # Verify
        expected = (
            '=IF(\'ICE\'!A2>0,"ICE: "&\'ICE\'!A2&", ","")'
            '&IF(\'HOP\'!A4>0,"HOP: "&\'HOP\'!A4&", ","")')
        self.assertEqual(expected, print_refs)

    def test_get_other_print_references_basic_land(self):
        # Setup
        mtgjson.update_models(self.session, self.mtg_data, True)
        self.session.commit()
        forests = self.session.query(models.CardPrinting).filter(
            models.CardPrinting.card.has(name='Forest')).all()
        name_to_prints = {'Forest': forests}
        lea_forest = self.session.query(
            models.CardPrinting).filter_by(multiverseid=288).first()

        # Execute
        print_refs = mtgxlsx.get_other_print_references(
            lea_forest, name_to_prints)

        # Verify
        self.assertIsNone(print_refs)

    def test_create_cards_sheet(self):
        # Setup
        mtgjson.update_models(self.session, self.mtg_data, True)
        self.session.commit()
        forest1 = self.session.query(
            models.CardPrinting).filter_by(multiverseid=2746).first()
        forest2 = self.session.query(
            models.CardPrinting).filter_by(multiverseid=2747).first()
        forest3 = self.session.query(
            models.CardPrinting).filter_by(multiverseid=2748).first()
        forest1.counts['copies'] = 1
        forest2.counts['foils'] = 2
        forest3.counts['copies'] = 3
        forest3.counts['foils'] = 4
        self.session.commit()
        printings = self.session.query(models.CardPrinting)
        name_to_prints = collections.defaultdict(list)
        for printing in printings:
            name_to_prints[printing.card.name].append(printing)
        ice_age = self.session.query(models.CardSet).filter_by(code='ICE').one()
        book = openpyxl.Workbook()
        sheet = book.create_sheet()

        # Execute
        mtgxlsx.create_cards_sheet(sheet, ice_age, name_to_prints)

        # Verify
        rows = [[cell.value for cell in row] for row in sheet.rows]
        # pylint: disable=line-too-long
        expected = [
            ['have', 'name', 'multiverseid', 'number', 'artist', 'copies', 'foils', 'others'],
            ['=F2+G2', 'Dark Ritual', 2444, None, 'Justin Hampton', None, None, mock.ANY],
            ['=F3+G3', 'Forest', 2746, None, 'Pat Morrissey', 1, None, mock.ANY],
            ['=F4+G4', 'Forest', 2747, None, 'Pat Morrissey', None, 2, mock.ANY],
            ['=F5+G5', 'Forest', 2748, None, 'Pat Morrissey', 3, 4, mock.ANY],
            ['=F6+G6', 'Snow-Covered Forest', 2749, None, 'Pat Morrissey', None, None, mock.ANY],
        ]
        # pylint: enable=line-too-long
        self.assertEqual(expected, rows)

    def test_dump_workbook(self):
        # Setup
        mtgjson.update_models(self.session, self.mtg_data, True)
        self.session.commit()

        # Execute
        book = mtgxlsx.dump_workbook(self.session)

        # Verify
        expected = [
            'Sets',
            'LEA',
            'ICE',
            'HML',
            'S00',
            'pMGD',
            'HOP',
            'ARC',
            'ISD',
            'PC2',
            'VMA',
        ]
        self.assertEqual(expected, book.sheetnames)

    def test_worksheet_row_reader(self):
        # Setup
        book = openpyxl.Workbook()
        sheet = book.create_sheet()
        sheet.title = 'ICE'
        sheet_data = [
            ['name', 'multiverseid', 'number', 'artist', 'copies', 'foils'],
            ['Forest', 2746, None, 'Pat Morrissey', 1, None],
            ['Forest', 2747, None, 'Pat Morrissey', None, 2],
            ['Forest', 2748, None, 'Pat Morrissey', 3, 4],
            ['Snow-Covered Forest', 2749, None, 'Pat Morrissey', None, None],
        ]
        for row in sheet_data:
            sheet.append(row)

        # Execute
        row_dicts = mtgxlsx.worksheet_row_reader(sheet)

        # Verify
        expected = [
            # pylint: disable=line-too-long
            {'name': 'Forest', 'artist': 'Pat Morrissey', 'multiverseid': 2746, 'number': None, 'copies': 1, 'foils': None},
            {'name': 'Forest', 'artist': 'Pat Morrissey', 'multiverseid': 2747, 'number': None, 'copies': None, 'foils': 2},
            {'name': 'Forest', 'artist': 'Pat Morrissey', 'multiverseid': 2748, 'number': None, 'copies': 3, 'foils': 4},
            {'name': 'Snow-Covered Forest', 'artist': 'Pat Morrissey', 'multiverseid': 2749, 'number': None, 'copies': None, 'foils': None},
            # pylint: enable=line-too-long
        ]
        self.assertEqual(expected, list(row_dicts))

    def test_workbook_row_reader(self):
        # Setup
        book = openpyxl.Workbook()
        sheet = book.create_sheet()
        sheet.title = 'ICE'
        sheet_data = [
            ['name', 'multiverseid', 'number', 'artist', 'copies', 'foils'],
            ['Forest', 2746, None, 'Pat Morrissey', 1, None],
            ['Forest', 2747, None, 'Pat Morrissey', None, 2],
            ['Forest', 2748, None, 'Pat Morrissey', 3, 4],
            ['Snow-Covered Forest', 2749, None, 'Pat Morrissey', None, None],
        ]
        for row in sheet_data:
            sheet.append(row)
        known_sets = {'ICE'}

        # Execute
        row_dicts = mtgxlsx.workbook_row_reader(book, known_sets)

        # Verify
        expected = [
            # pylint: disable=line-too-long
            {'set': 'ICE', 'name': 'Forest', 'artist': 'Pat Morrissey', 'multiverseid': 2746, 'number': None, 'copies': 1, 'foils': None},
            {'set': 'ICE', 'name': 'Forest', 'artist': 'Pat Morrissey', 'multiverseid': 2747, 'number': None, 'copies': None, 'foils': 2},
            {'set': 'ICE', 'name': 'Forest', 'artist': 'Pat Morrissey', 'multiverseid': 2748, 'number': None, 'copies': 3, 'foils': 4},
            {'set': 'ICE', 'name': 'Snow-Covered Forest', 'artist': 'Pat Morrissey', 'multiverseid': 2749, 'number': None, 'copies': None, 'foils': None},
            # pylint: enable=line-too-long
        ]
        self.assertEqual(expected, list(row_dicts))

    def test_workbook_row_reader_invalid_set(self):
        # Setup
        book = openpyxl.Workbook()
        sheet = book.create_sheet()
        sheet.title = 'garbage'
        sheet.append(['jabberwocky', 'gobbledy', 'goop'])
        known_sets = {'ICE'}

        # Execute
        row_dicts = mtgxlsx.workbook_row_reader(book, known_sets)

        # Verify
        self.assertFalse(list(row_dicts))

    def test_read_workbook_counts(self):
        # Setup
        mtgjson.update_models(self.session, self.mtg_data, True)
        self.session.commit()
        forest1 = self.session.query(
            models.CardPrinting).filter_by(multiverseid=2746).first()
        forest2 = self.session.query(
            models.CardPrinting).filter_by(multiverseid=2747).first()
        forest3 = self.session.query(
            models.CardPrinting).filter_by(multiverseid=2748).first()
        forest4 = self.session.query(
            models.CardPrinting).filter_by(multiverseid=2749).first()
        forest4.counts['copies'] = 2
        forest4.counts['foils'] = 3
        self.session.commit()
        book = openpyxl.Workbook()
        sets_sheet = book.create_sheet()
        sets_sheet.title = 'Sets'
        sets_sheet.append(['jabberwocky', 'gobbledy', 'goop'])

        cards_sheet = book.create_sheet()
        cards_sheet.title = 'ICE'
        sheet_data = [
            ['name', 'multiverseid', 'number', 'artist', 'copies', 'foils'],
            ['Forest', 2746, None, 'Pat Morrissey', 1, None],
            ['Forest', 2747, None, 'Pat Morrissey', None, 2],
            ['Forest', 2748, None, 'Pat Morrissey', 3, 4],
            ['Snow-Covered Forest', 2749, None, 'Pat Morrissey', None, None],
        ]
        for row in sheet_data:
            cards_sheet.append(row)

        # Execute
        mtgxlsx.read_workbook_counts(self.session, book)
        self.session.commit()

        # Verify
        self.assertEqual({'copies': 1}, forest1.counts)
        self.assertEqual({'foils': 2}, forest2.counts)
        self.assertEqual({'copies': 3, 'foils': 4}, forest3.counts)
        self.assertFalse(forest4.counts)
