"""Tests for mtg_ssm.serialization.xlsx."""

from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple
from uuid import UUID

import openpyxl
import pytest
from syrupy.assertion import SnapshotAssertion

from mtg_ssm.containers.bundles import ScryfallDataSet
from mtg_ssm.containers.collection import MagicCollection
from mtg_ssm.containers.counts import CountType, ScryfallCardCount
from mtg_ssm.containers.indexes import Oracle
from mtg_ssm.serialization import xlsx


@pytest.fixture(scope="session")
def oracle(scryfall_data: ScryfallDataSet) -> Oracle:
    """Oracle fixture."""
    accepted_sets = {
        "lea",
        "fem",
        "s00",
        "ice",
        "hop",
        "mbs",
        "pmbs",
        "chk",
        "bok",
        "sok",
        "neo",
        "pneo",
    }
    scryfall_data2 = ScryfallDataSet(
        sets=[s for s in scryfall_data.sets if s.code in accepted_sets],
        cards=[c for c in scryfall_data.cards if c.set in accepted_sets],
        migrations=[],
    )
    return Oracle(scryfall_data2)


def test_create_all_sets(snapshot: SnapshotAssertion, oracle: Oracle) -> None:
    book = openpyxl.Workbook()
    sheet = book.create_sheet()
    xlsx.create_all_sets(sheet, oracle.index)
    assert [[cell.value for cell in row] for row in sheet.rows] == snapshot(name=str(sheet.title))


def test_create_haverefs(oracle: Oracle) -> None:
    fem_thallids = [c for c in oracle.index.name_to_cards["Thallid"] if c.set == "fem"]
    fem_thallids.sort(key=lambda c: c.collector_number)
    haverefs = xlsx.create_haverefs(oracle.index, "fem", fem_thallids)
    assert haverefs == "SUM('FEM'!A2:A5)"


@pytest.mark.parametrize(
    ("name", "exclude_sets", "expected"),
    [
        ("Forest", None, None),
        ("Rhox", None, "=IF('S00'!A2>0,\"S00:\"&'S00'!A2,\"\")"),
        ("Rhox", {"s00"}, None),
        (
            "Dark Ritual",
            {"lea"},
            '=_xlfn.TEXTJOIN(", ",1,IF(\'ICE\'!A2>0,"ICE:"&\'ICE\'!A2,""),IF(\'HOP\'!A3>0,"HOP:"&\'HOP\'!A3,""))',
        ),
        ("Dark Ritual", {"lea", "ice"}, "=IF('HOP'!A3>0,\"HOP:\"&'HOP'!A3,\"\")"),
        (
            "Thallid",
            None,
            "=IF(SUM('FEM'!A2:A5)>0,\"FEM:\"&SUM('FEM'!A2:A5),\"\")",
        ),
    ],
)
def test_get_references(
    oracle: Oracle, name: str, exclude_sets: Optional[Set[str]], expected: str
) -> None:
    print_refs = xlsx.get_references(oracle.index, name, exclude_sets=exclude_sets)
    assert print_refs == expected


def test_create_all_cards_sheet(snapshot: SnapshotAssertion, oracle: Oracle) -> None:
    book = openpyxl.Workbook()
    sheet = book.create_sheet()
    xlsx.create_all_cards(sheet, oracle.index)
    assert [[cell.value for cell in row] for row in sheet.rows] == snapshot(name=str(sheet.title))


def test_create_set_sheet(snapshot: SnapshotAssertion, oracle: Oracle) -> None:
    card_counts: ScryfallCardCount = {
        UUID("fbdcbd97-90a9-45ea-94f6-2a1c6faaf965"): {CountType.NONFOIL: 1},
        UUID("b346b784-7bde-49d0-bfa9-56236cbe19d9"): {CountType.FOIL: 2},
        UUID("768c4d8f-5700-4f0a-9ff2-58422aeb1dac"): {
            CountType.NONFOIL: 3,
            CountType.FOIL: 4,
        },
    }
    collection = MagicCollection(oracle=oracle, counts=card_counts)
    book = openpyxl.Workbook()
    sheet = book.create_sheet()
    xlsx.create_set_sheet(sheet, collection, "ice")
    assert [[cell.value for cell in row] for row in sheet.rows] == snapshot(name=str(sheet.title))


def test_write(snapshot: SnapshotAssertion, oracle: Oracle, tmp_path: Path) -> None:
    xlsx_path = tmp_path / "outfile.xlsx"
    card_counts: ScryfallCardCount = {
        UUID("5d5f3f57-410f-4ee2-b93c-f5051a068828"): {
            CountType.NONFOIL: 7,
            CountType.FOIL: 12,
        }
    }
    collection = MagicCollection(oracle=oracle, counts=card_counts)

    serializer = xlsx.XlsxDialect()
    serializer.write(xlsx_path, collection)

    workbook = openpyxl.load_workbook(filename=xlsx_path)
    for i, sheet in enumerate(workbook.worksheets):
        assert [[cell.value for cell in row] for row in sheet.rows] == snapshot(
            name=f"{i:02d} - {sheet.title}"
        )


@pytest.mark.parametrize(
    ("sheets_and_rows", "skip_sheets", "expected"),
    [
        pytest.param(
            [("MySheet", [["A", "B", "C"], ["1", "B", "=5+7"]])],
            None,
            [{"set": "MySheet", "A": "1", "B": "B", "C": "=5+7"}],
        ),
        pytest.param(
            [
                ("MySheet", [["A", "B", "C"], ["1", "B", "=5+7"]]),
                (
                    "OtherSheet",
                    [["D", "E", "F"], ["D1", "E1", "F1"], ["D2", "E2", "F2"]],
                ),
            ],
            None,
            [
                {"set": "MySheet", "A": "1", "B": "B", "C": "=5+7"},
                {"set": "OtherSheet", "D": "D1", "E": "E1", "F": "F1"},
                {"set": "OtherSheet", "D": "D2", "E": "E2", "F": "F2"},
            ],
        ),
        pytest.param(
            [
                ("MySheet", [["A", "B", "C"], ["1", "B", "=5+7"]]),
                (
                    "OtherSheet",
                    [["D", "E", "F"], ["D1", "E1", "F1"], ["D2", "E2", "F2"]],
                ),
            ],
            {"OtherSheet"},
            [{"set": "MySheet", "A": "1", "B": "B", "C": "=5+7"}],
        ),
    ],
)
def test_rows_from_workbook(
    sheets_and_rows: List[Tuple[str, List[List[str]]]],
    skip_sheets: Optional[Set[str]],
    expected: List[Dict[str, str]],
) -> None:
    workbook = openpyxl.Workbook()
    for sheet, rows in sheets_and_rows:
        worksheet = workbook.create_sheet(title=sheet)
        for row in rows:
            worksheet.append(row)
    del workbook["Sheet"]
    assert list(xlsx.rows_for_workbook(workbook, skip_sheets=skip_sheets)) == expected


def test_read_from_file(oracle: Oracle, tmp_path: Path) -> None:
    xlsx_path = tmp_path / "infile.xlsx"
    serializer = xlsx.XlsxDialect()
    workbook = openpyxl.Workbook()
    sheet = workbook["Sheet"]
    sheet.title = "S00"
    sheet.append(["scryfall_id", "nonfoil", "foil"])
    sheet.append(["5d5f3f57-410f-4ee2-b93c-f5051a068828", 3, 7])
    workbook.save(xlsx_path)
    collection = serializer.read(xlsx_path, oracle)
    assert collection.counts == {
        UUID("5d5f3f57-410f-4ee2-b93c-f5051a068828"): {
            CountType.NONFOIL: 3,
            CountType.FOIL: 7,
        }
    }
