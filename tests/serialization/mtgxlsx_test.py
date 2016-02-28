"""Tests for mtg_ssm.mtgxlsx"""

import datetime
from unittest import mock

import openpyxl

from mtg_ssm.mtg import collection
from mtg_ssm.mtg import models
from mtg_ssm.serialization import mtgxlsx

from tests.mtgjson import mtgjson_testcase


class MtgXlsxTest(mtgjson_testcase.MtgJsonTestCase):

    def setUp(self):
        super().setUp()
        self.collection = collection.Collection(
            self.mtg_data, include_online_only=True)

    def test_create_sets_sheet(self):
        # Setup
        card_sets = self.collection.code_to_card_set.values()
        card_sets = sorted(card_sets, key=lambda cset: cset.release_date)
        book = openpyxl.Workbook()
        sheet = book.create_sheet()

        # Execute
        mtgxlsx.create_sets_sheet(sheet, card_sets)

        # Verify
        rows = [[cell.value for cell in row] for row in sheet.rows]
        expected = [
            # pylint: disable=line-too-long
            ['code', 'name', 'release', 'block', 'type', 'cards', 'unique', 'playsets', 'count'],
            ['Total', None, None, None, None, '=SUM(F3:F65535)', '=SUM(G3:G65535)', '=SUM(H3:H65535)', '=SUM(I3:I63353)'],
            ['LEA', 'Limited Edition Alpha', datetime.datetime(1993, 8, 5), None, 'core', 4, '=COUNTIF(\'LEA\'!A:A,">0")', '=COUNTIF(\'LEA\'!A:A,">=4")', "=SUM('LEA'!A:A)"],
            ['FEM', 'Fallen Empires', datetime.datetime(1994, 11, 1), None, 'expansion', 4, '=COUNTIF(\'FEM\'!A:A,">0")', '=COUNTIF(\'FEM\'!A:A,">=4")', "=SUM('FEM'!A:A)"],
            ['pMEI', 'Media Inserts', datetime.datetime(1995, 1, 1), None, 'promo', 1, '=COUNTIF(\'pMEI\'!A:A,">0")', '=COUNTIF(\'pMEI\'!A:A,">=4")', "=SUM('pMEI'!A:A)"],
            ['ICE', 'Ice Age', datetime.datetime(1995, 6, 1), 'Ice Age', 'expansion', 5, '=COUNTIF(\'ICE\'!A:A,">0")', '=COUNTIF(\'ICE\'!A:A,">=4")', "=SUM('ICE'!A:A)"],
            ['HML', 'Homelands', datetime.datetime(1995, 10, 1), None, 'expansion', 2, '=COUNTIF(\'HML\'!A:A,">0")', '=COUNTIF(\'HML\'!A:A,">=4")', "=SUM('HML'!A:A)"],
            ['S00', 'Starter 2000', datetime.datetime(2000, 4, 1), None, 'starter', 1, '=COUNTIF(\'S00\'!A:A,">0")', '=COUNTIF(\'S00\'!A:A,">=4")', "=SUM('S00'!A:A)"],
            ['PLS', 'Planeshift', datetime.datetime(2001, 2, 5), 'Invasion', 'expansion', 2, '=COUNTIF(\'PLS\'!A:A,">0")', '=COUNTIF(\'PLS\'!A:A,">=4")', "=SUM('PLS'!A:A)"],
            ['pMGD', 'Magic Game Day', datetime.datetime(2007, 7, 14), None, 'promo', 1, '=COUNTIF(\'pMGD\'!A:A,">0")', '=COUNTIF(\'pMGD\'!A:A,">=4")', "=SUM('pMGD'!A:A)"],
            ['HOP', 'Planechase', datetime.datetime(2009, 9, 4), None, 'planechase', 3, '=COUNTIF(\'HOP\'!A:A,">0")', '=COUNTIF(\'HOP\'!A:A,">=4")', "=SUM('HOP'!A:A)"],
            ['ARC', 'Archenemy', datetime.datetime(2010, 6, 18), None, 'archenemy', 2, '=COUNTIF(\'ARC\'!A:A,">0")', '=COUNTIF(\'ARC\'!A:A,">=4")', "=SUM('ARC'!A:A)"],
            ['ISD', 'Innistrad', datetime.datetime(2011, 9, 30), 'Innistrad', 'expansion', 6, '=COUNTIF(\'ISD\'!A:A,">0")', '=COUNTIF(\'ISD\'!A:A,">=4")', "=SUM('ISD'!A:A)"],
            ['PC2', 'Planechase 2012 Edition', datetime.datetime(2012, 6, 1), None, 'planechase', 4, '=COUNTIF(\'PC2\'!A:A,">0")', '=COUNTIF(\'PC2\'!A:A,">=4")', "=SUM('PC2'!A:A)"],
            ['MMA', 'Modern Masters', datetime.datetime(2013, 6, 7, 0, 0), None, 'reprint', 1, '=COUNTIF(\'MMA\'!A:A,">0")', '=COUNTIF(\'MMA\'!A:A,">=4")', "=SUM('MMA'!A:A)"],
            ['VMA', 'Vintage Masters', datetime.datetime(2014, 6, 16, 0, 0), None, 'masters', 1, '=COUNTIF(\'VMA\'!A:A,">0")', '=COUNTIF(\'VMA\'!A:A,">=4")', "=SUM('VMA'!A:A)"],
        ]
        self.assertEqual(expected, rows)

    def test_split_into_consecutives(self):
        inlist = []
        expected = []
        self.assertEqual(expected, mtgxlsx.split_into_consecutives(inlist))

        inlist = [1]
        expected = [[1]]
        self.assertEqual(expected, mtgxlsx.split_into_consecutives(inlist))

        inlist = [1, 2, 3]
        expected = [[1, 2, 3]]
        self.assertEqual(expected, mtgxlsx.split_into_consecutives(inlist))

        inlist = [1, 2, 3, 5, 6, 8, 9, 10, 11]
        expected = [[1, 2, 3], [5, 6], [8, 9, 10, 11]]
        self.assertEqual(expected, mtgxlsx.split_into_consecutives(inlist))

        inlist = [4, 3, 7, 1]
        expected = [[1], [3, 4], [7]]
        self.assertEqual(expected, mtgxlsx.split_into_consecutives(inlist))

    def test_create_haveref_sum(self):
        setcode = 'ABC'

        rownums = [1]
        expected = "'ABC'!A1"
        self.assertEqual(expected, mtgxlsx.create_haveref_sum(setcode, rownums))

        rownums = [3, 4, 5, 6]
        expected = "SUM('ABC'!A3:A6)"
        self.assertEqual(expected, mtgxlsx.create_haveref_sum(setcode, rownums))

        rownums = [3, 4, 5, 8, 10]
        expected = "SUM('ABC'!A3:A5)+'ABC'!A8+'ABC'!A10"
        self.assertEqual(expected, mtgxlsx.create_haveref_sum(setcode, rownums))

    def test_get_refs_multiple_sets(self):
        # Setup
        lea_dark_rit = self.collection.id_to_printing[
            'fff0b8e8fea06ee1ac5c35f048a0a459b1222673']

        # Execute
        print_refs = mtgxlsx.get_other_print_references(lea_dark_rit)

        # Verify
        expected = (
            '=IF(\'ICE\'!A2>0,"ICE: "&\'ICE\'!A2&", ","")'
            '&IF(\'HOP\'!A4>0,"HOP: "&\'HOP\'!A4&", ","")')
        self.assertEqual(expected, print_refs)

    def test_get_refs_multiple_variants(self):
        # Setup
        mma_thallid = self.collection.id_to_printing[
            'fc46a4b72d216117a352f59217a84d0baeaaacb7']

        # Execute
        print_refs = mtgxlsx.get_other_print_references(mma_thallid)

        # Verify
        expected = (
            '=IF(SUM(\'FEM\'!A2:A5)>0,"FEM: "&SUM(\'FEM\'!A2:A5)&", ","")')
        self.assertEqual(expected, print_refs)

    def test_get_refs_basic_land(self):
        # Setup
        lea_forest = self.collection.id_to_printing[
            '5ede9781b0c5d157c28a15c3153a455d7d6180fa']

        # Execute
        print_refs = mtgxlsx.get_other_print_references(lea_forest)

        # Verify
        self.assertIsNone(print_refs)

    def test_create_cards_sheet(self):
        # Setup
        forest1 = self.collection.id_to_printing[
            '676a1f5b64dc03bbb3876840c3ff2ba2c16f99cb']
        forest2 = self.collection.id_to_printing[
            'd0a4414893bc2f9bd3beea2f8f2693635ef926a4']
        forest3 = self.collection.id_to_printing[
            'c78d2da78c68c558b1adc734b3f164e885407ffc']
        forest1.counts[models.CountTypes.copies] = 1
        forest2.counts[models.CountTypes.foils] = 2
        forest3.counts[models.CountTypes.copies] = 3
        forest3.counts[models.CountTypes.foils] = 4

        ice_age = self.collection.code_to_card_set['ICE']
        book = openpyxl.Workbook()
        sheet = book.create_sheet()

        # Execute
        mtgxlsx.create_cards_sheet(sheet, ice_age)

        # Verify
        rows = [[cell.value for cell in row] for row in sheet.rows]
        expected = [
            # pylint: disable=line-too-long
            ['have', 'name', 'id', 'multiverseid', 'number', 'artist', 'copies', 'foils', 'others'],
            ['=G2+H2', 'Dark Ritual', '2fab0ea29e3bbe8bfbc981a4c8163f3e7d267853', 2444, None, 'Justin Hampton', None, None, mock.ANY],
            ['=G3+H3', 'Forest', '676a1f5b64dc03bbb3876840c3ff2ba2c16f99cb', 2746, None, 'Pat Morrissey', 1, None, mock.ANY],
            ['=G4+H4', 'Forest', 'd0a4414893bc2f9bd3beea2f8f2693635ef926a4', 2747, None, 'Pat Morrissey', None, 2, mock.ANY],
            ['=G5+H5', 'Forest', 'c78d2da78c68c558b1adc734b3f164e885407ffc', 2748, None, 'Pat Morrissey', 3, 4, mock.ANY],
            ['=G6+H6', 'Snow-Covered Forest', '5e9f08498a9343b1954103e493da2586be0fe394', 2749, None, 'Pat Morrissey', None, None, mock.ANY],
        ]
        self.assertEqual(expected, rows)

    def test_dump_workbook(self):
        # Execute
        book = mtgxlsx.dump_workbook(self.collection)

        # Verify
        expected = [
            'Sets',
            'LEA',
            'FEM',
            'pMEI',
            'ICE',
            'HML',
            'S00',
            'PLS',
            'pMGD',
            'HOP',
            'ARC',
            'ISD',
            'PC2',
            'MMA',
            'VMA',
        ]
        self.assertEqual(expected, book.sheetnames)

    def test_worksheet_row_reader(self):
        # Setup
        book = openpyxl.Workbook()
        sheet = book.create_sheet()
        sheet.title = 'ICE'
        sheet_data = [
            ['name', 'multiverseid', 'number', 'artist', 'copies', 'foils'],
            ['Forest', 2746, None, 'Pat Morrissey', 1, None],
            ['Forest', 2747, None, 'Pat Morrissey', None, 2],
            ['Forest', 2748, None, 'Pat Morrissey', 3, 4],
            ['Snow-Covered Forest', 2749, None, 'Pat Morrissey', None, None],
        ]
        for row in sheet_data:
            sheet.append(row)

        # Execute
        row_dicts = mtgxlsx.worksheet_row_reader(sheet)

        # Verify
        expected = [
            # pylint: disable=line-too-long
            {'name': 'Forest', 'artist': 'Pat Morrissey', 'multiverseid': 2746, 'number': None, 'copies': 1, 'foils': None},
            {'name': 'Forest', 'artist': 'Pat Morrissey', 'multiverseid': 2747, 'number': None, 'copies': None, 'foils': 2},
            {'name': 'Forest', 'artist': 'Pat Morrissey', 'multiverseid': 2748, 'number': None, 'copies': 3, 'foils': 4},
            {'name': 'Snow-Covered Forest', 'artist': 'Pat Morrissey', 'multiverseid': 2749, 'number': None, 'copies': None, 'foils': None},
        ]
        self.assertEqual(expected, list(row_dicts))

    def test_workbook_row_reader(self):
        # Setup
        book = openpyxl.Workbook()
        sheet = book.create_sheet()
        sheet.title = 'ICE'
        sheet_data = [
            ['name', 'multiverseid', 'number', 'artist', 'copies', 'foils'],
            ['Forest', 2746, None, 'Pat Morrissey', 1, None],
            ['Forest', 2747, None, 'Pat Morrissey', None, 2],
            ['Forest', 2748, None, 'Pat Morrissey', 3, 4],
            ['Snow-Covered Forest', 2749, None, 'Pat Morrissey', None, None],
        ]
        for row in sheet_data:
            sheet.append(row)
        known_sets = {'ICE'}

        # Execute
        row_dicts = mtgxlsx.workbook_row_reader(book, known_sets)

        # Verify
        expected = [
            # pylint: disable=line-too-long
            {'set': 'ICE', 'name': 'Forest', 'artist': 'Pat Morrissey', 'multiverseid': 2746, 'number': None, 'copies': 1, 'foils': None},
            {'set': 'ICE', 'name': 'Forest', 'artist': 'Pat Morrissey', 'multiverseid': 2747, 'number': None, 'copies': None, 'foils': 2},
            {'set': 'ICE', 'name': 'Forest', 'artist': 'Pat Morrissey', 'multiverseid': 2748, 'number': None, 'copies': 3, 'foils': 4},
            {'set': 'ICE', 'name': 'Snow-Covered Forest', 'artist': 'Pat Morrissey', 'multiverseid': 2749, 'number': None, 'copies': None, 'foils': None},
        ]
        self.assertEqual(expected, list(row_dicts))

    def test_row_reader_invalid_set(self):
        # Setup
        book = openpyxl.Workbook()
        sheet = book.create_sheet()
        sheet.title = 'garbage'
        sheet.append(['jabberwocky', 'gobbledy', 'goop'])
        known_sets = {'ICE'}

        # Execute
        row_dicts = mtgxlsx.workbook_row_reader(book, known_sets)

        # Verify
        self.assertFalse(list(row_dicts))

    def test_read_workbook_counts(self):
        # Setup
        forest1 = self.collection.id_to_printing[
            '676a1f5b64dc03bbb3876840c3ff2ba2c16f99cb']
        forest2 = self.collection.id_to_printing[
            'd0a4414893bc2f9bd3beea2f8f2693635ef926a4']
        forest3 = self.collection.id_to_printing[
            'c78d2da78c68c558b1adc734b3f164e885407ffc']
        forest4 = self.collection.id_to_printing[
            '5e9f08498a9343b1954103e493da2586be0fe394']
        forest4.counts[models.CountTypes.copies] = 2
        forest4.counts[models.CountTypes.foils] = 3

        book = openpyxl.Workbook()
        sets_sheet = book.create_sheet()
        sets_sheet.title = 'Sets'
        sets_sheet.append(['jabberwocky', 'gobbledy', 'goop'])

        cards_sheet = book.create_sheet()
        cards_sheet.title = 'ICE'
        sheet_data = [
            ['name', 'multiverseid', 'number', 'artist', 'copies', 'foils'],
            ['Forest', 2746, None, 'Pat Morrissey', 1, None],
            ['Forest', 2747, None, 'Pat Morrissey', None, 2],
            ['Forest', 2748, None, 'Pat Morrissey', 3, 4],
            ['Snow-Covered Forest', 2749, None, 'Pat Morrissey', None, None],
        ]
        for row in sheet_data:
            cards_sheet.append(row)

        # Execute
        mtgxlsx.read_workbook_counts(self.collection, book)

        # Verify
        self.assertEqual({models.CountTypes.copies: 1}, forest1.counts)
        self.assertEqual({models.CountTypes.foils: 2}, forest2.counts)
        self.assertEqual(
            {models.CountTypes.copies: 3, models.CountTypes.foils: 4},
            forest3.counts)
        self.assertEqual(
            {models.CountTypes.copies: 2, models.CountTypes.foils: 3},
            forest4.counts)
